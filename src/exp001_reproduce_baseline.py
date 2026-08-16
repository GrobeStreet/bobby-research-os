#!/usr/bin/env python3
import csv
import hashlib
import json
import statistics
import urllib.request
from collections import OrderedDict, defaultdict
from pathlib import Path

import openpyxl

DATA_URL = 'https://pasteur.epa.gov/uploads/10.23719/1531811/GAC%20matrix%20impact%20paper%20data_ScID_D-pnwm.xlsx'
DATA_SHA256 = '91d941d114f7df43578b9cf45c8fbb1d29caf5fa802401982afcd3588ab3bb5b'
DATA_FILE = Path('data/raw/pfas_gac_matrix.xlsx')
OUT_CSV = Path('outputs/exp001_baseline_summary.csv')
OUT_JSON = Path('outputs/exp001_baseline_result.json')


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def ensure_data() -> str:
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    if not DATA_FILE.exists():
        urllib.request.urlretrieve(DATA_URL, DATA_FILE)
    digest = sha256(DATA_FILE)
    if digest != DATA_SHA256:
        raise RuntimeError(f'Dataset checksum mismatch: expected {DATA_SHA256}, got {digest}')
    return digest


def summarize_sheet(ws, condition_col: int, percent_col: int, pfas_col: int):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    conditions = []
    for row in rows:
        if len(row) <= max(condition_col, percent_col, pfas_col):
            continue
        if row[condition_col] is None:
            continue
        condition = str(row[condition_col]).strip()
        if condition not in conditions:
            conditions.append(condition)

    summary = OrderedDict()
    for condition in conditions:
        vals = []
        pfas_vals = defaultdict(list)
        for r in rows:
            if len(r) <= max(condition_col, percent_col, pfas_col):
                continue
            if r[condition_col] is None or str(r[condition_col]).strip() != condition or r[percent_col] is None:
                continue
            try:
                value = float(r[percent_col])
            except (TypeError, ValueError):
                continue
            vals.append(value)
            pfas_vals[str(r[pfas_col]).strip()].append(value)
        if vals:
            summary[condition] = {
                'n': len(vals),
                'median_percent_change': statistics.median(vals),
                'mean_percent_change': statistics.mean(vals),
                'min_percent_change': min(vals),
                'max_percent_change': max(vals),
                'pfas_medians': {p: statistics.median(v) for p, v in sorted(pfas_vals.items())},
            }
    return summary


def main() -> int:
    digest = ensure_data()
    wb = openpyxl.load_workbook(DATA_FILE, read_only=True, data_only=True)

    required = {'Fig 4', 'Fig S4'}
    missing = required - set(wb.sheetnames)
    if missing:
        raise RuntimeError(f'Missing expected sheets: {sorted(missing)}')

    fig4 = summarize_sheet(wb['Fig 4'], condition_col=1, percent_col=2, pfas_col=0)
    figs4 = summarize_sheet(wb['Fig S4'], condition_col=1, percent_col=3, pfas_col=0)

    expected_conditions = [
        '9PFAS, 10 mM NaHCO3',
        'pH 9',
        '5 mM sodium bicarbonate',
        '2 mM sodium bicarbonate',
        'calcium sulfate',
        'NOM (co-load)',
        'NOM (co-load) + calcium sulfate',
    ]
    absent = [c for c in expected_conditions if c not in figs4]
    if absent:
        raise RuntimeError(f'Expected Fig S4 conditions not parsed: {absent}')

    checks = OrderedDict([
        ('baseline_is_zero', abs(figs4['9PFAS, 10 mM NaHCO3']['median_percent_change']) < 1e-9),
        ('pH9_is_modest_negative', figs4['pH 9']['median_percent_change'] < 0),
        ('5mM_bicarbonate_increases_capacity', figs4['5 mM sodium bicarbonate']['median_percent_change'] > 0),
        ('2mM_bicarbonate_increases_capacity', figs4['2 mM sodium bicarbonate']['median_percent_change'] > 0),
        ('calcium_sulfate_increases_capacity', figs4['calcium sulfate']['median_percent_change'] > 0),
        ('NOM_coload_reduces_capacity', figs4['NOM (co-load)']['median_percent_change'] < 0),
        ('calcium_sulfate_partially_offsets_NOM', figs4['NOM (co-load) + calcium sulfate']['median_percent_change'] > figs4['NOM (co-load)']['median_percent_change']),
    ])

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open('w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['sheet', 'condition', 'n', 'median_percent_change', 'mean_percent_change', 'min_percent_change', 'max_percent_change'])
        for sheet_name, summary in [('Fig 4', fig4), ('Fig S4', figs4)]:
            for condition, stats in summary.items():
                w.writerow([sheet_name, condition, stats['n'], round(stats['median_percent_change'], 6), round(stats['mean_percent_change'], 6), round(stats['min_percent_change'], 6), round(stats['max_percent_change'], 6)])

    result = {
        'experiment_id': 'EXP-001',
        'dataset': {'doi': '10.23719/1531811', 'sha256': digest, 'url': DATA_URL},
        'checks': checks,
        'status': 'PASS' if all(checks.values()) else 'FAIL',
        'fig_s4_key_medians_percent_change': {k: round(v['median_percent_change'], 3) for k, v in figs4.items()},
        'interpretation_boundary': 'Directional reproduction only. Not a full-scale treatment-performance model or operational guidance.',
    }
    OUT_JSON.write_text(json.dumps(result, indent=2) + '\n', encoding='utf-8')
    print(json.dumps(result, indent=2))
    return 0 if result['status'] == 'PASS' else 1


if __name__ == '__main__':
    raise SystemExit(main())
